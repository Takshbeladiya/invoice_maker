import streamlit as st
import math
import base64
from datetime import datetime
from io import BytesIO
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.units import inch
from tinydb import TinyDB, Query
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

# --- DB SETUP (UNCHANGED) ---
def setup_database():
    """Initializes TinyDB, ensures default products are present, and reads all product names."""
    db = TinyDB('db.json')
    product_table = db.table('products')
    db.table('invoices') 
    
    Product = Query()
    default_products = ['sunscreen', 'zebra', 'solid_blackout', 'curtain']
    
    for name in default_products:
        if not product_table.search(Product.name == name):
            product_table.insert({'name': name, 'fields': ['price', 'profit_ratio']})
            
    all_product_names = [p['name'] for p in product_table.all() if 'name' in p and isinstance(p['name'], str)]
    return db, all_product_names

# --- DYNAMICALLY INITIALIZE SESSION STATE (SIMPLIFIED) ---
def initialize_session_state(products):
    """Initializes session state, dynamically adding keys for all available products."""
    if 'blinds_data' not in st.session_state:
        st.session_state.blinds_data = []
    if 'show_add_form' not in st.session_state:
        st.session_state.show_add_form = False
    if 'prefill_data' not in st.session_state:
        st.session_state.prefill_data = None
    if 'editing_blind_id' not in st.session_state:
        st.session_state.editing_blind_id = None

    if 'excel_report_data' not in st.session_state:
        st.session_state.excel_report_data = None

    if 'pdf_report_data' not in st.session_state:
        st.session_state.pdf_report_data = None

    if 'pdf_report_data_no_amount' not in st.session_state:
        st.session_state.pdf_report_data_no_amount = None
    if 'pdf_report_data_with_amount' not in st.session_state:
        st.session_state.pdf_report_data_with_amount = None

    if 'motor_price' not in st.session_state:
        st.session_state.motor_price = 0.0
    if 'motor_quantity' not in st.session_state:
        st.session_state.motor_quantity = 0
    if 'motor_shipping_price' not in st.session_state:
        st.session_state.motor_shipping_price = 0.0

    if 'pricing' not in st.session_state:
        st.session_state.pricing = {}
    if 'profit_percentages' not in st.session_state:
        st.session_state.profit_percentages = {}

    defaults = {
        'sunscreen': {'price': 2.1, 'ratio': 0.3, 'profit': 0.7},
        'zebra': {'price': 2.6, 'ratio': 0.3, 'profit': 0.7},
        'solid_blackout': {'price': 1.7, 'ratio': 0.3, 'profit': 0.7},
        'curtain': {'price': 1.9, 'ratio': 0.3, 'profit': 0.7},
    }

    for product in products:
        price_key = f"{product}_price"
        ratio_key = f"{product}_profit_ratio"
        profit_key = f"profit_{product}"
        universal_ratio_key = f"universal_ratio_{product}"

        if price_key not in st.session_state.pricing:
            st.session_state.pricing[price_key] = defaults.get(product, {}).get('price', 1.0)
        if ratio_key not in st.session_state.pricing:
            st.session_state.pricing[ratio_key] = defaults.get(product, {}).get('ratio', 0.3)
        if product not in st.session_state.profit_percentages:
            st.session_state.profit_percentages[product] = defaults.get(product, {}).get('profit', 0.7)
        if universal_ratio_key not in st.session_state:
            st.session_state[universal_ratio_key] = defaults.get(product, {}).get('ratio', 0.3)
        if profit_key not in st.session_state:
            st.session_state[profit_key] = defaults.get(product, {}).get('profit', 0.7)

# --- CALCULATION LOGIC - SHIPPING COST FIXED ---
def calculate_blind_costs(width, height, total_blinds, mount, pricing, shipping_rate, selected_products, total_pieces=None):
    """Calculates costs for all selected products using the updated formulas.
    
    Shipping Cost Formula:
    (((width_in_cm * 13 * 13) / 4850) * 10) / shipping_rate * (undivided * total)
    Where:
    - undivided = total_blinds
    - total = total_pieces
    """
    if total_pieces is None:
        total_pieces = total_blinds
    
    total_sqft_final = (width * height / 144) * (total_blinds * total_pieces)
    
    number_of_splits = 1
    actual_total_pieces = total_pieces
    shipping_cost_final = 0
    if shipping_rate > 0:
        width_in_cm = (width + 2) * 2.5
        shipping_cost_base = (((width_in_cm * 13 * 13) / 4850) * 10) / shipping_rate
        # Multiply by (undivided * total) = (total_blinds * total_pieces)
        shipping_cost_final = shipping_cost_base * (total_blinds * total_pieces)

    product_costs = {}
    for name in selected_products.keys(): 
        is_selected = selected_products.get(name, False)
        cost_key = f"{name}_cost"
        if is_selected:
            price = pricing.get(f'{name}_price', 0)
            ratio = pricing.get(f'{name}_profit_ratio', 1)
            if ratio > 0:
                cost = total_sqft_final * (price / ratio)
            else:
                cost = 0
            product_costs[cost_key] = cost
        else:
            product_costs[cost_key] = 0

    results = {
        'total_sqft': total_sqft_final, 'shipping_cost': shipping_cost_final,
        'actual_total_pieces': actual_total_pieces, 'number_of_splits': number_of_splits,
    }
    results.update(product_costs)
    return results

# --- Recalculation Function (SIMPLIFIED) ---
def recalculate_all_blinds():
    """Loops through all blinds and recalculates their costs based on their currently stored data."""
    updated_blinds_data = []
    for blind in st.session_state.blinds_data:
        total_pieces = blind.get('total_pieces', blind.get('total_blinds', 1))
        updated_costs = calculate_blind_costs(
            width=blind['width'], height=blind['height'], total_blinds=blind['total_blinds'],
            mount=blind['mount'], pricing=blind.get('pricing', {}), shipping_rate=blind['shipping_rate'],
            selected_products=blind['selected_products'], total_pieces=total_pieces
        )
        updated_blind = blind.copy()
        updated_blind.update(updated_costs)
        updated_blinds_data.append(updated_blind)
    st.session_state.blinds_data = updated_blinds_data

# --- NEW: Bulk Update Function ---
def bulk_update_ratios(active_products):
    """Updates the profit ratio for all blinds of a specific product type."""
    for product_name in active_products:
        new_ratio_key = f"universal_ratio_{product_name}"
        new_ratio = st.session_state.get(new_ratio_key)

        if new_ratio is not None:
            for blind in st.session_state.blinds_data:
                if blind.get('selected_products', {}).get(product_name, False):
                    blind['pricing'][f"{product_name}_profit_ratio"] = new_ratio


# --- INVOICE GENERATION FUNCTION (UNCHANGED) ---
def generate_invoice_pdf_no_amount(invoice_data):
    """Generates a landscape PDF invoice with an itemized breakdown WITHOUT costs."""
    LOGO_BASE64 = "" 

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    story = []

    if LOGO_BASE64:
        # ... (logo logic remains the same)
        pass

    header_data = [
        [Paragraph('<b>Interior Shin</b><br/>123 Decor St, Window City, 12345<br/>contact@interiorshin.com<br/>+1 940-594-8904', styles['Normal']), 
         Paragraph(f'<b>INVOICE</b><br/>Date: {datetime.now().strftime("%B %d, %Y")}', ParagraphStyle(name='Right', parent=styles['Normal'], alignment=2))]
    ]
    header_table = Table(header_data, colWidths=[doc.width/2.0]*2)
    header_table.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP')]))
    story.append(header_table)
    story.append(Spacer(1, 0.25*inch))
    story.append(Paragraph("<b>Itemized Breakdown</b>", styles['h3']))

    has_splits = any(b.get('number_of_splits', 1) > 1 for b in invoice_data['blinds_data'])
    headers = ['ID', 'Desc', 'W', 'H', 'Window', 'Blind Per Window', 'Mount']
    if has_splits: 
        headers.insert(6, 'Split')
    
    table_data = [[Paragraph(f'<b>{h}</b>', styles['Normal']) for h in headers]]

    for blind in invoice_data['blinds_data']:
        row = [
            blind['id'], 
            Paragraph(blind.get('description', 'N/A'), styles['Normal']), 
            f"{blind['width']}\"", f"{blind['height']}\"",
            blind['total_blinds'], blind.get('actual_total_pieces'), blind.get('mount')
        ]
        if has_splits:
            number_of_splits = blind.get('number_of_splits', 1)
            split_display = f"{blind['width']}/{number_of_splits}" if number_of_splits > 1 else "-"
            row.insert(6, split_display)
        table_data.append(row)

    if invoice_data.get('motor_quantity', 0) > 0:
        motor_qty = invoice_data['motor_quantity']
        motor_row = ['-', Paragraph('Motor', styles['Normal']), '', '', motor_qty, motor_qty, '-']
        if has_splits:
            motor_row.insert(6, '-') 
        table_data.append(motor_row)

    item_table = Table(table_data, repeatRows=1)
    item_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey), 
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('ALIGN', (0,0), (-1,-1), 'CENTER'), 
        ('ALIGN', (1,1), (1,-1), 'LEFT'),
    ]))
    story.append(item_table)
    
    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes

def generate_invoice_pdf_with_amount(invoice_data):
    """Generates a landscape PDF invoice with a full cost breakdown."""
    LOGO_BASE64 = ""

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    styles = getSampleStyleSheet()
    story = []

    # --- Header ---
    header_data = [
        [Paragraph('<b>Interior Shin</b><br/>123 Decor St, Window City, 12345<br/>contact@interiorshin.com<br/>+1 940-594-8904', styles['Normal']),
         Paragraph(f'<b>INVOICE</b><br/>Date: {datetime.now().strftime("%B %d, %Y")}', ParagraphStyle(name='Right', parent=styles['Normal'], alignment=2))]
    ]
    header_table = Table(header_data, colWidths=[doc.width/2.0]*2)
    header_table.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP')]))
    story.append(header_table)
    story.append(Spacer(1, 0.2*inch))
    story.append(Paragraph("<b>Itemized Breakdown</b>", styles['h3']))

    # --- Main Table with Costs ---
    active_products = invoice_data['active_products']
    has_splits = any(b.get('number_of_splits', 1) > 1 for b in invoice_data['blinds_data'])
    
    headers = ['ID', 'Desc', 'W', 'H', 'Window', 'Blind Per Window']
    if has_splits:
        headers.append('Split')
    headers.append('Mount')
    headers.extend([p.replace('_', ' ').capitalize() for p in active_products])
    headers.append('Shipping')
    
    table_data = [[Paragraph(f'<b>{h}</b>', styles['Normal']) for h in headers]]

    # --- Blinds Rows ---
    for blind in invoice_data['blinds_data']:
        row = [
            blind['id'], Paragraph(blind.get('description', 'N/A'), styles['Normal']),
            f"{blind['width']}\"", f"{blind['height']}\"",
            blind['total_blinds'], blind.get('actual_total_pieces')
        ]
        if has_splits:
            number_of_splits = blind.get('number_of_splits', 1)
            row.append(f"{blind['width']}/{number_of_splits}" if number_of_splits > 1 else "-")
        row.append(blind.get('mount'))
        
        for p in active_products:
            cost = blind.get(f'{p}_cost', 0)
            row.append(f"${cost:.2f}" if cost > 0 else "-")
        
        row.append(f"${blind.get('shipping_cost', 0):.2f}")
        table_data.append(row)

    # --- Motor Row ---
    if invoice_data.get('motor_quantity', 0) > 0:
        motor_row = ['-', Paragraph('Motor', styles['Normal']), '', '', invoice_data['motor_quantity'], invoice_data['motor_quantity']]
        if has_splits:
            motor_row.append('-')
        motor_row.append('-') # Mount

        motor_qty = invoice_data['motor_quantity']
        motor_price_per_unit = invoice_data['motor_price']
        motor_shipping_per_unit = invoice_data['motor_shipping_price']


        for i, p in enumerate(active_products):
            motor_row.append(f"${motor_price_per_unit:.2f}" if i == 0 else "-")
        motor_row.append(f"${motor_shipping_per_unit:.2f}")
        table_data.append(motor_row)

    item_table = Table(table_data, repeatRows=1, hAlign='LEFT')
    item_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('ALIGN', (1,1), (1,-1), 'LEFT'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
    ]))
    story.append(item_table)
    story.append(Spacer(1, 0.25*inch))
    
    # --- Summary Section ---
    summary_data = []
    bold_style = ParagraphStyle(name='BoldRight', parent=styles['Normal'], alignment=2, fontName='Helvetica-Bold')
    right_align_style = ParagraphStyle(name='Right', parent=styles['Normal'], alignment=2)
    
    # Product Sub-Totals and Shipping
    for p in active_products:
        p_name = p.replace('_', ' ').capitalize()
        summary_data.append([Paragraph(f"{p_name} Sub-Total:", styles['Normal']), Paragraph(f"${invoice_data['sub_totals'].get(p, 0):.2f}", right_align_style)])
        summary_data.append([Paragraph(f"{p_name} Est. Shipping:", styles['Normal']), Paragraph(f"${invoice_data['shipping_totals'].get(p, 0):.2f}", right_align_style)])

    # Overall Totals
    summary_data.append([Paragraph("<b>Overall Sub-Total:</b>", styles['Normal']), Paragraph(f"${invoice_data['overall_sub_total']:.2f}", bold_style)])
    summary_data.append([Paragraph("<b>Total Estimated Shipping:</b>", styles['Normal']), Paragraph(f"${invoice_data['overall_shipping_total']:.2f}", bold_style)])
    
    # Grand Total (Bill Total)
    bill_total_p1 = Paragraph("<b>Bill Total:</b>", styles['Normal'])
    bill_total_p2 = Paragraph(f"${invoice_data['bill_total']:.2f}", bold_style)
    summary_data.append([bill_total_p1, bill_total_p2])

    summary_table = Table(summary_data, colWidths=[2.5*inch, 1*inch], hAlign='RIGHT')
    summary_table.setStyle(TableStyle([
        ('BOX', (0, -1), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
    ]))
    story.append(summary_table)
    
    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes

# --- EXCEL GENERATION FUNCTION (RESIZE COLUMN REMOVED) ---
# --- UPDATED EXCEL GENERATION FUNCTION ---
def generate_excel_report(blinds_data, session_state, active_products):
    """
    Generates an Excel report with customer details, blind table, and complete summary.
    Uses latest calculation methods: Product Cost = Price * Sq Ft
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Blinds Calculation"

    # --- STYLES ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    input_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    summary_label_font = Font(bold=True)
    bill_total_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    customer_label_font = Font(bold=True)

    # --- CUSTOMER DETAILS SECTION (A1-A3 and C1) ---
    ws.cell(row=1, column=1, value="Customer Name:")
    ws.cell(row=1, column=1).font = customer_label_font
    
    ws.cell(row=2, column=1, value="Address:")
    ws.cell(row=2, column=1).font = customer_label_font
    
    ws.cell(row=3, column=1, value="Phone No:")
    ws.cell(row=3, column=1).font = customer_label_font
    
    date_cell = ws.cell(row=1, column=3, value=f"Date: {datetime.now().strftime('%B %d, %Y')}")
    date_cell.font = customer_label_font
    
    # --- BLIND TABLE SECTION (starting from A5) ---
    table_start_row = 5
    
    headers = ['ID', 'Desc', 'W', 'H', 'Window', 'Blind Per Window', 'Mount', 'Total Sq Ft']
    for product in active_products:
        headers.append(f"{product.replace('_', ' ').capitalize()} Cost")
    headers.extend(['Shipping Cost', 'Line Total'])
    
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=table_start_row, column=i, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # --- WRITE BLIND DATA ROWS ---
    row_num = table_start_row + 1
    for blind in sorted(blinds_data, key=lambda x: x['id']):
        ws.cell(row=row_num, column=1, value=blind['id'])
        ws.cell(row=row_num, column=2, value=blind.get('description', 'N/A'))
        ws.cell(row=row_num, column=3, value=blind['width']).fill = input_fill
        ws.cell(row=row_num, column=4, value=blind['height']).fill = input_fill
        ws.cell(row=row_num, column=5, value=blind['total_blinds'])
        ws.cell(row=row_num, column=6, value=blind.get('actual_total_pieces', blind['total_blinds']))
        ws.cell(row=row_num, column=7, value=blind['mount'])

        w_cell, h_cell, num_cell_undivided, num_cell_total = f'C{row_num}', f'D{row_num}', f'E{row_num}', f'F{row_num}'

        # Sq Ft Formula
        sqft_formula = f"=IFERROR(ROUND((({w_cell}*{h_cell})/144)*({num_cell_undivided}*{num_cell_total}), 2), 0)"
        sqft_cell_obj = ws.cell(row=row_num, column=8, value=sqft_formula)
        sqft_cell_obj.number_format = '0.00'
        sqft_cell = f'H{row_num}'
        
        # Product Cost Formulas (using new method: Price * Sq Ft)
        product_cost_cells = []
        col_idx = 9
        for p in active_products:
            blind_pricing = blind.get('pricing', {})
            price = blind_pricing.get(f'{p}_price', session_state.pricing.get(f'{p}_price', 0))
            ratio = blind_pricing.get(f'{p}_profit_ratio', session_state.pricing.get(f'{p}_profit_ratio', 1))
            
            # NEW FORMULA: Product Cost = Price * Sq Ft
            p_formula = f"=IFERROR(IF({ratio}>0, ROUND({sqft_cell}*({price}/{ratio}), 2), 0), 0)"
            
            # Only show cost if product is selected
            if not blind.get('selected_products', {}).get(p, False): 
                p_formula = 0
            
            p_cell_obj = ws.cell(row=row_num, column=col_idx, value=p_formula)
            p_cell_obj.number_format = '$#,##0.00'
            product_cost_cells.append(f'{get_column_letter(col_idx)}{row_num}')
            col_idx += 1

        # Shipping Cost Formula
        shipping_rate = blind.get('shipping_rate', 0.9)
        shipping_formula = f"=IF({shipping_rate}>0, ROUND((((((({w_cell}+2)*2.5)*13*13)/4850)*10)/{shipping_rate})*({num_cell_total}*{num_cell_undivided}), 2), 0)"
        shipping_cell_obj = ws.cell(row=row_num, column=col_idx, value=shipping_formula)
        shipping_cell_obj.number_format = '$#,##0.00'
        shipping_cell = f'{get_column_letter(col_idx)}{row_num}'
        col_idx += 1
        
        # Line Total Formula
        line_total_formula = f"=ROUND(SUM({','.join(product_cost_cells)}, {shipping_cell}), 2)"
        line_total_obj = ws.cell(row=row_num, column=col_idx, value=line_total_formula)
        line_total_obj.number_format = '$#,##0.00'
        
        row_num += 1

    # --- MOTOR ROW ---
    if session_state.motor_quantity > 0:
        ws.cell(row=row_num, column=2, value="Motor")
        ws.cell(row=row_num, column=5, value=session_state.motor_quantity)
        ws.cell(row=row_num, column=6, value=session_state.motor_quantity)
        
        first_product_col_idx = 9
        shipping_col_idx = 8 + len(active_products) + 1
        total_col_idx = shipping_col_idx + 1

        motor_price_cell_obj = ws.cell(row=row_num, column=first_product_col_idx, value=session_state.motor_price)
        motor_price_cell_obj.number_format = '$#,##0.00'
        
        motor_shipping_cell_obj = ws.cell(row=row_num, column=shipping_col_idx, value=session_state.motor_shipping_price)
        motor_shipping_cell_obj.number_format = '$#,##0.00'

        motor_qty_cell = f'F{row_num}'
        motor_price_cell = f'{get_column_letter(first_product_col_idx)}{row_num}'
        motor_shipping_cell = f'{get_column_letter(shipping_col_idx)}{row_num}'
        motor_total_formula = f"=ROUND((({motor_price_cell} + {motor_shipping_cell})*{motor_qty_cell}), 2)"
        motor_total_obj = ws.cell(row=row_num, column=total_col_idx, value=motor_total_formula)
        motor_total_obj.number_format = '$#,##0.00'
        
        row_num += 1

    last_data_row = row_num - 1
    last_blind_row = last_data_row - (1 if session_state.motor_quantity > 0 else 0)
    summary_start_row = row_num + 2
    
    # --- SUMMARY SECTION (NEW FORMAT) ---
    label_col, value_col = 2, 3
    current_summary_row = summary_start_row
    
    # # Product Cost Summary (using new calculation: Price * Sq Ft)
    # ws.cell(row=current_summary_row, column=label_col, value="PRODUCT COST SUMMARY").font = summary_label_font
    # current_summary_row += 1
    
    # for i, p in enumerate(active_products):
    #     product_col_letter = get_column_letter(9 + i)
    #     product_total_formula = f"=ROUND(SUM({product_col_letter}{table_start_row+1}:{product_col_letter}{last_blind_row}), 2)"
    #     p_clean = p.replace('_', ' ').capitalize()
    #     product_cost_summmary_saved_row = current_summary_row
    #     ws.cell(row=current_summary_row, column=label_col, value=f"{p_clean} Cost:")

    #     product_summary_cell = ws.cell(row=current_summary_row, column=value_col, value=product_total_formula)
    #     product_summary_cell.number_format = '"$"#,##0.00'
    #     current_summary_row += 1

    # Shipping Summary
    current_summary_row += 1
    ws.cell(row=current_summary_row, column=label_col, value="SHIPPING SUMMARY").font = summary_label_font
    current_summary_row += 1
    
    shipping_col_letter = get_column_letter(8 + len(active_products) + 1)
    shipping_total_formula = f"=ROUND(SUM({shipping_col_letter}{table_start_row+1}:{shipping_col_letter}{last_blind_row}), 2)"
    ws.cell(row=current_summary_row, column=label_col, value="Total Estimated Shipping:")
    shipping_summary_cell = ws.cell(row=current_summary_row, column=value_col, value=shipping_total_formula)
    shipping_summary_cell.number_format = '"$"#,##0.00'
    current_summary_row += 1

    # Profit Summary
    
    ws.cell(row=current_summary_row, column=label_col, value="PROFIT SUMMARY").font = summary_label_font
    current_summary_row += 1
    
    # Store cost and profit cell references
    product_cost_refs = {}    # product → cost cell coordinate
    product_profit_refs = []  # list of profit cell coordinates (for summing later)

    for i, p in enumerate(active_products):
        product_col_letter = get_column_letter(9 + i)
        p_clean = p.replace('_', ' ').capitalize()
        
        # Instead of summing the Excel column (which has / ratio),
        # calculate total cost using only price * sqft
        total_cost_direct = 0.0
        for blind in blinds_data:
            if not blind.get('selected_products', {}).get(p, False):
                continue
            sqft = blind.get('total_sqft', 0)
            price = blind.get('pricing', {}).get(f'{p}_price',
                        session_state.pricing.get(f'{p}_price', 0))
            total_cost_direct += price * sqft

        ws.cell(row=current_summary_row, column=label_col, value=f"{p_clean} Cost (price × sqft):")
        
        product_summary_cell = ws.cell(
            row=current_summary_row,
            column=value_col,
            value=round(total_cost_direct, 2)
        )
        product_summary_cell.number_format = '"$"#,##0.00'
        
        product_cost_refs[p] = product_summary_cell.coordinate
        
        current_summary_row += 1

    # Now write Profit lines and collect their cell references
    for i, p in enumerate(active_products):
        p_clean = p.replace('_', ' ').capitalize()
        cost_ref = product_cost_refs.get(p, "0")   # still useful for reference

        # Calculate total profit by looping over actual blinds data
        profit_value = 0.0
        for blind in blinds_data:
            if not blind.get('selected_products', {}).get(p, False):
                continue
            
            blind_cost = blind.get(f'{p}_cost', 0)
            ratio = blind.get('pricing', {}).get(f'{p}_profit_ratio', 0.3)  # fallback if missing
            
            if 0 < ratio < 1:
                profit_value += (1 - ratio) * blind_cost

        # Write the calculated value (static number, not formula)
        profit_cell = ws.cell(
            row=current_summary_row,
            column=value_col,
            value=round(profit_value, 2)
        )
        profit_cell.number_format = '"$"#,##0.00'

        # Label — you can customize the text
        ws.cell(
            row=current_summary_row,
            column=label_col,
            value=f"{p_clean} Profit (per blind ratio):"
        )

        # Still collect for net profit sum later
        product_profit_refs.append(profit_cell.coordinate)
        
        current_summary_row += 1

        # Total Calculations
    current_summary_row += 1
    ws.cell(row=current_summary_row, column=label_col, value="TOTALS").font = summary_label_font
    current_summary_row += 1

    # ────────────────────────────────────────────────
    # Overall Sub-Total (sum of all product costs)
    # ────────────────────────────────────────────────
    product_cost_cols = [get_column_letter(9 + i) for i in range(len(active_products))]
    overall_subtotal_formula = f"=ROUND(SUM({','.join([f'{col}{table_start_row+1}:{col}{last_blind_row}' for col in product_cost_cols])}), 2)"

    ws.cell(row=current_summary_row, column=label_col, value="Overall Sub-Total:")
    overall_subtotal_cell = ws.cell(
        row=current_summary_row,
        column=value_col,
        value=overall_subtotal_formula
    )
    overall_subtotal_cell.number_format = '"$"#,##0.00'
    overall_subtotal_ref = overall_subtotal_cell.coordinate     # e.g. C22
    current_summary_row += 1

    # ────────────────────────────────────────────────
    # Total Estimated Shipping (only blinds)
    # ────────────────────────────────────────────────
    ws.cell(row=current_summary_row, column=label_col, value="Total Estimated Shipping:")
    total_shipping_cell = ws.cell(
        row=current_summary_row,
        column=value_col,
        value=shipping_total_formula
    )
    total_shipping_cell.number_format = '"$"#,##0.00'
    total_shipping_ref = total_shipping_cell.coordinate         # e.g. C23
    current_summary_row += 1

    # ────────────────────────────────────────────────
    # Motor Cost (includes motor shipping)
    # ────────────────────────────────────────────────
    motor_cost_formula = f"={session_state.motor_quantity}*({session_state.motor_price}+{session_state.motor_shipping_price})"

    ws.cell(row=current_summary_row, column=label_col, value="Motor Cost:")
    motor_cost_cell = ws.cell(
        row=current_summary_row,
        column=value_col,
        value=motor_cost_formula
    )
    motor_cost_cell.number_format = '"$"#,##0.00'
    motor_cost_ref = motor_cost_cell.coordinate                 # e.g. C24
    current_summary_row += 1

    # ────────────────────────────────────────────────
    # Net Profit = sum of individual product profit cells
    # ────────────────────────────────────────────────
    ws.cell(row=current_summary_row, column=label_col, value="Net Profit:")
    
    if product_profit_refs:
        net_profit_formula = f"=ROUND(SUM({','.join(product_profit_refs)}), 2)"
    else:
        net_profit_formula = "0"

    net_profit_cell = ws.cell(
        row=current_summary_row,
        column=value_col,
        value=net_profit_formula
    )
    net_profit_cell.number_format = '"$"#,##0.00'
    net_profit_ref = net_profit_cell.coordinate                  # e.g. C25
    current_summary_row += 1

    # ────────────────────────────────────────────────
    # BILL TOTAL = Overall Sub-Total + Net Profit + Total Est. Shipping + Motor Cost
    # ────────────────────────────────────────────────
    bill_total_label_cell = ws.cell(
        row=current_summary_row,
        column=label_col,
        value="BILL TOTAL:"
    )
    bill_total_label_cell.font = summary_label_font
    bill_total_label_cell.fill = bill_total_fill

    # Use the actual cell coordinates we just created
    bill_total_formula = f"=ROUND({overall_subtotal_ref} + {motor_cost_ref} + {total_shipping_ref}, 2)"

    bill_total_value_cell = ws.cell(
        row=current_summary_row,
        column=value_col,
        value=bill_total_formula
    )
    bill_total_value_cell.font = summary_label_font
    bill_total_value_cell.fill = bill_total_fill
    bill_total_value_cell.number_format = '"$"#,##0.00'
    bill_total_ref = bill_total_value_cell.coordinate            # e.g. C26
    current_summary_row += 1

    # ────────────────────────────────────────────────
    # Net Profit %
    # ────────────────────────────────────────────────
    net_profit_percent_formula = f"=IF({bill_total_ref}>0, ROUND({net_profit_ref}/{bill_total_ref}*100, 2), 0)"

    ws.cell(row=current_summary_row, column=label_col, value="Net Profit %:")
    net_profit_percent_cell = ws.cell(
        row=current_summary_row,
        column=value_col,
        value=net_profit_percent_formula
    )
    net_profit_percent_cell.number_format = '0.00"%"'
    



    # --- COLUMN WIDTHS ---
    for col_idx_num in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx_num)].width = 15
    ws.column_dimensions['B'].width = 30

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# --- ADD/EDIT FORM (RESIZE CHECKBOX REMOVED) ---
def add_blind_form(available_products):
    is_editing = st.session_state.editing_blind_id is not None
    
    defaults = st.session_state.prefill_data if is_editing and st.session_state.prefill_data else {
        'description': "", 'width': 36.0, 'height': 95.5, 'total_blinds': 2, 'total_pieces': 2, 'mount': "Inside Mt",
        'shipping_rate': 0.9, 'selected_products': {}, 'pricing': st.session_state.pricing
    }

    if 'form_product_selection' not in st.session_state:
        st.session_state.form_product_selection = [name for name, selected in defaults.get('selected_products', {}).items() if selected]

    form_title = "✏️ Edit Blind" if is_editing else "p Add New Blind"
    st.subheader(form_title)
    
    st.multiselect("Products", options=available_products, default=st.session_state.form_product_selection, key="form_product_selection")

    with st.form("add_edit_blind_form"):
        description = st.text_input("Description", value=defaults.get('description', ''), placeholder="e.g., Living room window")
        col1, col2 = st.columns(2)
        with col1:
            width = st.number_input("Width (inches)", min_value=1.0, value=float(defaults.get('width')), step=0.5)
            height = st.number_input("Height (inches)", min_value=1.0, value=float(defaults.get('height')), step=0.5)
        with col2:
            total_blinds = st.number_input("Number of Windows", min_value=1, value=int(defaults.get('total_blinds')))
            total_pieces = st.number_input("Blind Per Window", min_value=1, value=int(defaults.get('total_pieces', defaults.get('total_blinds'))), help="Total pieces used to calculate shipping cost")
        
        col3, col4 = st.columns(2)
        with col3:
            mount = st.selectbox("Mount Type", ["Inside Mt", "Outside Mt"], index=["Inside Mt", "Outside Mt"].index(defaults.get('mount', "Inside Mt")))
        with col4:
            shipping_rate = st.number_input("Shipping Rate", min_value=0.1, value=float(defaults.get('shipping_rate')), step=0.1)
        
        selected_product_names = st.session_state.form_product_selection
        pricing_inputs = {}
        if selected_product_names:
            st.markdown("**Price Settings**")
            price_cols = st.columns(len(selected_product_names))
            for i, name in enumerate(selected_product_names):
                with price_cols[i]:
                    st.markdown(f"*{name.replace('_', ' ').capitalize()}*")
                    price_key, ratio_key = f"{name}_price", f"{name}_profit_ratio"
                    default_price = defaults.get('pricing', {}).get(price_key, st.session_state.pricing.get(price_key, 0))
                    default_ratio = defaults.get('pricing', {}).get(ratio_key, st.session_state.pricing.get(ratio_key, 0.3))
                    price_str = st.text_input("Price", value=str(default_price), key=f"price_{name}")
                    ratio_str = st.text_input("Profit Ratio", value=str(default_ratio), key=f"ratio_{name}")
                    pricing_inputs[name] = {'price_str': price_str, 'ratio_str': ratio_str}

        submit_col, cancel_col = st.columns(2)
        with submit_col:
            submitted = st.form_submit_button("✅ " + ("Update Blind" if is_editing else "Add Blind to Table"), type="primary")
        with cancel_col:
            if st.form_submit_button("❌ Cancel"):
                st.session_state.show_add_form = False
                st.session_state.prefill_data, st.session_state.editing_blind_id = None, None
                del st.session_state.form_product_selection
                st.rerun()

        if submitted:
            custom_pricing = {}
            try:
                for name, inputs in pricing_inputs.items():
                    custom_pricing[f"{name}_price"] = float(inputs['price_str'])
                    custom_pricing[f"{name}_profit_ratio"] = float(inputs['ratio_str'])
            except ValueError:
                st.error("Text not allowed. Please enter valid numbers for prices and profit ratios.")
                st.stop()

            selected_products = {name: (name in selected_product_names) for name in available_products}
            final_pricing = st.session_state.pricing.copy()
            final_pricing.update(custom_pricing)

            costs = calculate_blind_costs(width, height, total_blinds, mount, final_pricing, shipping_rate, selected_products, total_pieces)
            blind_id = st.session_state.editing_blind_id or ((max(b['id'] for b in st.session_state.blinds_data) + 1) if st.session_state.blinds_data else 1)
            
            if is_editing:
                st.session_state.blinds_data = [b for b in st.session_state.blinds_data if b['id'] != blind_id]

            new_blind_data = {
                'id': blind_id, 'description': description, 'width': width, 'height': height,
                'total_blinds': total_blinds, 'total_pieces': total_pieces, 'mount': mount, 'shipping_rate': shipping_rate,
                'selected_products': selected_products, 'pricing': final_pricing, **costs
            }
            st.session_state.blinds_data.append(new_blind_data)

            st.success(f"Blind #{blind_id} processed successfully!")
            st.session_state.show_add_form = False
            st.session_state.prefill_data, st.session_state.editing_blind_id = None, None
            del st.session_state.form_product_selection
            st.rerun()

# --- DISPLAY TABLE (DATAFRAME VERSION) ---
def display_blinds_table():
    if not st.session_state.blinds_data:
        return

    st.subheader("📋 Blinds Summary")
    
    active_products = sorted(list(set(
        key.replace('_cost', '') 
        for blind in st.session_state.blinds_data 
        for key in blind 
        if key.endswith('_cost') and blind[key] > 0 and key != 'shipping_cost'
    )))
    
    # --- Build DataFrame ---
    table_data = []
    for blind in sorted(st.session_state.blinds_data, key=lambda x: x['id']):
        row = {
            'ID': blind['id'],
            'Description': blind.get('description', 'N/A'),
            'Width (")': blind['width'],
            'Height (")': blind['height'],
            'Windows': blind['total_blinds'],
            'Blind Per Window': blind.get('total_pieces', blind.get('actual_total_pieces', blind['total_blinds'])),
            'Mount': blind['mount'],
            'Sq Ft': round(blind['total_sqft'], 2),
        }
        
        # Add product costs
        for product_name in active_products:
            cost = blind.get(f'{product_name}_cost', 0)
            col_name = f"{product_name.replace('_', ' ').capitalize()}"
            row[col_name] = f"${cost:.2f}" if cost > 0 else "-"
        
        # Add shipping
        row['Shipping ($)'] = f"${blind.get('shipping_cost', 0):.2f}"
        
        # Calculate line total
        line_total = sum(
            blind.get(f'{p}_cost', 0) 
            for p in active_products
        ) + blind.get('shipping_cost', 0)
        row['Line Total ($)'] = f"${line_total:.2f}"
        
        table_data.append(row)
    
    df = pd.DataFrame(table_data)
    
    # --- Display Table with styling ---
    st.dataframe(
        df, 
        use_container_width=True, 
        hide_index=True,
        column_config={
            'ID': st.column_config.NumberColumn(format="%d"),
            'Width (")': st.column_config.NumberColumn(format="%.1f"),
            'Height (")': st.column_config.NumberColumn(format="%.1f"),
            'Windows': st.column_config.NumberColumn(format="%d"),
            'Blind Per Window': st.column_config.NumberColumn(format="%d"),
            'Sq Ft': st.column_config.NumberColumn(format="%.2f"),
        }
    )
    
    # --- Action Buttons Below Table ---
    st.markdown("**Actions:**")
    for blind in sorted(st.session_state.blinds_data, key=lambda x: x['id']):
        col1, col2, col3 = st.columns([8, 1, 1])
        with col1:
            st.text(f"Blind #{blind['id']} - {blind.get('description', 'N/A')}")
        with col2:
            if st.button("✏️ Edit", key=f"edit_{blind['id']}", help=f"Edit Blind {blind['id']}"):
                st.session_state.prefill_data, st.session_state.editing_blind_id = blind, blind['id']
                st.session_state.show_add_form = True
                st.rerun()
        with col3:
            if st.button("🗑️ Delete", key=f"delete_{blind['id']}", help=f"Delete Blind {blind['id']}"):
                st.session_state.blinds_data = [b for b in st.session_state.blinds_data if b['id'] != blind['id']]
                st.success(f"Blind #{blind['id']} deleted!")
                recalculate_all_blinds()
                st.rerun()

def calculate_product_costs_detailed(blinds_data, p_name):
    product_cost = 0.0
    product_sqft = 0.0
    
    for blind in blinds_data:
        blind_sqft = blind.get('total_sqft', 0)
        selected_products = blind.get('selected_products', {})
        blind_pricing = blind.get('pricing', {})
        
        if selected_products.get(p_name, False):
            product_price = blind_pricing.get(f'{p_name}_price', 0)
            cost_for_blind = product_price * blind_sqft
            
            product_cost += cost_for_blind
            product_sqft += blind_sqft
    
    return product_cost


# --- MAIN APPLICATION FLOW (UPDATED)---
def main():
    st.set_page_config(page_title="Blinds Cost Calculator", layout="wide")
    st.title("🪟 Blinds Cost Calculator")
    
    db, available_products = setup_database()
    initialize_session_state(available_products)

    # --- Sidebar for Load/Delete (UNCHANGED) ---
    st.sidebar.title("Invoice History")
    invoices_table = db.table('invoices')
    saved_invoices = invoices_table.all()
    invoice_names = ["--- Start New Invoice ---"] + sorted([inv.get('name', f"Untitled_{inv.doc_id}") for inv in saved_invoices], reverse=True)
    
    selected_invoice_name = st.sidebar.selectbox("Load Saved Invoice", invoice_names, key="invoice_selector")
    
    col1, col2 = st.sidebar.columns(2)
    with col1:
        if st.button("Load Invoice"):
            if selected_invoice_name != "--- Start New Invoice ---":
                Invoice = Query()
                invoice_data = invoices_table.get(Invoice.name == selected_invoice_name)
                if invoice_data:
                    for key, value in invoice_data.get('session_data', {}).items():
                        st.session_state[key] = value
                    st.success(f"Invoice '{selected_invoice_name}' loaded successfully!")
                    st.rerun()
            else:
                st.sidebar.info("Select an invoice to load.")
    with col2:
        if st.button("Delete Invoice", help="Deletes the selected invoice permanently."):
            if selected_invoice_name != "--- Start New Invoice ---":
                Invoice = Query()
                invoices_table.remove(Invoice.name == selected_invoice_name)
                st.sidebar.success(f"Deleted '{selected_invoice_name}'.")
                st.session_state.invoice_selector = "--- Start New Invoice ---"
                st.rerun()
            else:
                st.sidebar.warning("Select an invoice to delete.")

    st.markdown("---")
    if st.session_state.show_add_form: 
        add_blind_form(available_products)
    else:
        if st.button("p Add New Blind", type="primary"):
            st.session_state.show_add_form = True
            st.session_state.prefill_data, st.session_state.editing_blind_id = None, None
            if 'form_product_selection' in st.session_state:
                del st.session_state.form_product_selection
            st.rerun()
    st.markdown("---")
    
    display_blinds_table()
    
    if st.session_state.blinds_data:
        st.markdown("---")
        active_products_for_settings = sorted(list(set(
            p for blind in st.session_state.blinds_data
            for p, selected in blind.get('selected_products', {}).items() if selected
        )))

    #     st.subheader("Profit Settings")
    #     if not active_products_for_settings:
    #         st.info("Add a blind and select products to set their final profit percentages.")
    #     else:
    #         profit_cols = st.columns(len(active_products_for_settings))
    #         for i, p_name in enumerate(active_products_for_settings):
    #             with profit_cols[i]:
    #                 st.number_input(f"{p_name.replace('_',' ').capitalize()} Profit %", min_value=0, max_value=100, 
    # step=1, format="%d", key=f"profit_{p_name}")
    #     st.markdown("---")

        st.subheader("⚙️ Motor Settings")
        motor_cols = st.columns(3)
        with motor_cols[0]:
            st.number_input("Motor Price ($)", min_value=0.0, key="motor_price", step=10.0, format="%.2f")
        with motor_cols[1]:
            st.number_input("Motor Quantity", min_value=0, key="motor_quantity", step=1)
        with motor_cols[2]:
            st.number_input("Motor Shipping ($)", min_value=0.0, key="motor_shipping_price", step=5.0, format="%.2f")
        st.markdown("---")

        st.subheader("💰 Cost Summary")

        if st.button("🔄 Apply Changes", key="apply_changes_btn", use_container_width=False):
            st.session_state.apply_changes = True
            st.rerun()

        sub_totals = {p: sum(b.get(f'{p}_cost', 0) for b in st.session_state.blinds_data) for p in available_products}
        shipping_totals = {p: sum(b.get('shipping_cost', 0) for b in st.session_state.blinds_data if b.get('selected_products', {}).get(p)) for p in available_products}
        
        profit_totals = {p: (sub_totals[p] * st.session_state.get(f"profit_{p}", 0) / 100) for p in available_products}
        total_motor_cost = st.session_state.motor_quantity * (st.session_state.motor_price + st.session_state.motor_shipping_price)
        
        net_profit = sum(profit_totals.values())
        total_sqft = sum(b.get('total_sqft', 0) for b in st.session_state.blinds_data) 

        # Change Needed
        total_pieces = sum([(b.get('total_blinds') * b.get('actual_total_pieces')) for b in st.session_state.blinds_data])
        
        overall_sub_total = sum(sub_totals.values())
        overall_shipping_total = sum(b.get('shipping_cost', 0) for b in st.session_state.blinds_data)

        

        active_products_summary = {p: v for p, v in sub_totals.items() if v > 0}
        num_summary_cols = len(active_products_summary) + 1 if active_products_summary else 1
        summary_cols = st.columns(num_summary_cols)

        i = 0
        bill_total_cal = []
        net_profit_tot = []
        for p_name, sub_total_val in active_products_summary.items():
            with summary_cols[i]:
                sub_total_calculation = calculate_product_costs_detailed(st.session_state.blinds_data, p_name)   
                

                st.metric(f"{p_name.replace('_',' ').capitalize()} Cost", f"${sub_total_calculation:.2f}")
                st.metric(f"{p_name.replace('_',' ').capitalize()} Est. Shipping", f"${shipping_totals[p_name]:.2f}")
                # # Each product Profit
                # cost_with_profit = ((sub_total_calculation * 100)/(100 - (st.session_state.get(f"profit_{p_name}", 0)))) 
                # print(st.session_state.get(f"profit_{p_name}", 0) * 10)
                # st.metric(f"{p_name.replace('_',' ').capitalize()} Profit", f"${cost_with_profit:.2f}")
                # bill_total_cal.append(cost_with_profit)
                # bill_total_cal.append(sub_total_calculation)
                # net_profit_tot.append(cost_with_profit)

                # New: sum over each blind using its OWN profit ratio
                product_profit = 0.0
                for blind in st.session_state.blinds_data:
                    if not blind.get('selected_products', {}).get(p_name, False):
                        continue
                    cost = blind.get(f"{p_name}_cost", 0)
                    ratio = blind.get('pricing', {}).get(f"{p_name}_profit_ratio", 0.3)  # fallback 0.3
                    if ratio <= 0 or ratio >= 1:
                        continue
                    # Profit = (1 - ratio) * cost
                    product_profit += (1 - ratio) * cost

                st.metric(
                    f"{p_name.replace('_',' ').capitalize()} Profit",
                    f"${product_profit:.2f}"
                )

                bill_total_cal.append(product_profit)      # ← keep this line
                bill_total_cal.append(sub_total_calculation)
                net_profit_tot.append(product_profit)      # ← keep this line

            i += 1
            
        # Change in 28 Feb 
        bill_total = sum(bill_total_cal) + overall_shipping_total + total_motor_cost

        with summary_cols[-1]:
            st.metric("Total Sq Ft", f"{total_sqft:.2f}")
            st.metric("Total Pieces", f"{total_pieces}")
            st.metric("Motor Cost", f"${total_motor_cost:.2f}")
            st.metric("💵 Bill Total", f"${bill_total:.2f}")
            st.metric("🎯 Net Profit", f"${sum(net_profit_tot):.2f}")
            # Change made final
            st.metric("Net Profit %", f"{(sum(net_profit_tot)/bill_total)*100:.2f}%")
        st.markdown("---")
        
        st.subheader("Actions")
        col1, col2 = st.columns(2)
        with col1:
            st.info("Generate and download a report of the current invoice.")
            
            # --- Generation Buttons ---
            gen_col1, gen_col2, gen_col3 = st.columns(3)
            with gen_col1:
                if st.button("📄 Gen. Excel"):
                    with st.spinner("Creating Excel report..."):
                        st.session_state.excel_report_data = generate_excel_report(
                            st.session_state.blinds_data, st.session_state, active_products_for_settings)
            with gen_col2:
                if st.button("📄 Gen. PDF (No Amount)"):
                     with st.spinner("Creating PDF invoice..."):
                        invoice_data = {
                            'blinds_data': st.session_state.blinds_data,
                            'motor_quantity': st.session_state.get('motor_quantity', 0)
                        }
                        st.session_state.pdf_report_data_no_amount = generate_invoice_pdf_no_amount(invoice_data)
            with gen_col3:
                if st.button("💵 Gen. PDF (With Amount)"):
                    with st.spinner("Creating detailed PDF invoice..."):
                        invoice_data = {
                            'blinds_data': st.session_state.blinds_data,
                            'active_products': active_products_for_settings,
                            'motor_quantity': st.session_state.get('motor_quantity', 0),
                            'motor_price': st.session_state.get('motor_price', 0),
                            'motor_shipping_price': st.session_state.get('motor_shipping_price', 0),
                            'sub_totals': sub_totals,
                            'shipping_totals': shipping_totals,
                            'overall_sub_total': overall_sub_total,
                            'overall_shipping_total': overall_shipping_total,
                            'bill_total': bill_total
                        }
                        st.session_state.pdf_report_data_with_amount = generate_invoice_pdf_with_amount(invoice_data)
            
            # --- Download Buttons (appear after generation) ---
            dl_col1, dl_col2, dl_col3 = st.columns(3)
            with dl_col1:
                if st.session_state.excel_report_data:
                    st.download_button(label="✅ Download Excel", data=st.session_state.excel_report_data,
                        file_name=f"blinds_report_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        on_click=lambda: st.session_state.update(excel_report_data=None))
            with dl_col2:
                if st.session_state.pdf_report_data_no_amount:
                    st.download_button(label="✅ Download PDF", data=st.session_state.pdf_report_data_no_amount,
                        file_name=f"invoice_summary_{datetime.now().strftime('%Y%m%d')}.pdf", mime='application/pdf',
                        on_click=lambda: st.session_state.update(pdf_report_data_no_amount=None))
            with dl_col3:
                if st.session_state.pdf_report_data_with_amount:
                    st.download_button(label="✅ Download PDF", data=st.session_state.pdf_report_data_with_amount,
                        file_name=f"invoice_detailed_{datetime.now().strftime('%Y%m%d')}.pdf", mime='application/pdf',
                        on_click=lambda: st.session_state.update(pdf_report_data_with_amount=None))
                    
        with col2:
            st.info("Save the current invoice to the database.")
            invoice_name = st.text_input("Invoice Name", placeholder="e.g., Smith Job - Phase 1")
            if st.button("💾 Save Current Invoice", type="primary"):
                if invoice_name:
                    keys_to_save = ['blinds_data', 'motor_price', 'motor_quantity', 'motor_shipping_price', 'pricing']
                    session_data_to_save = {key: st.session_state[key] for key in keys_to_save if key in st.session_state}
                    for key in st.session_state:
                        if key.startswith('profit_') or key.startswith('universal_ratio_'):
                            session_data_to_save[key] = st.session_state[key]
                    invoice_data = {"name": invoice_name, "saved_at": datetime.now().isoformat(), "session_data": session_data_to_save}
                    Invoice = Query()
                    invoices_table.upsert(invoice_data, Invoice.name == invoice_name)
                    st.success(f"Invoice '{invoice_name}' saved successfully!")
                    st.rerun()
                else:
                    st.warning("Please enter an invoice name before saving.")
        
        st.markdown("---")
        if st.button("🗑️ Clear All Blinds"):
            st.session_state.blinds_data = []
            st.session_state.motor_quantity = 0
            st.rerun()

    st.markdown("---")
    with st.expander("📐 View Calculation Formulas"):
        st.markdown("""
        **1. Core Blind Calculation (per line item):**
        - **Total Sq Ft** = `((Width * Height) / 144) * Number of Blinds`
        - **Product Sub-Total** = `Total Sq Ft * (Product Price / Profit Ratio)`
        
        **2. Shipping Cost Calculation (CORRECTED):**
        - `Shipping Cost Base` = `(((Width in CM * 13 * 13) / 4850) * 10) / Shipping Rate`
        - `Shipping Cost` = `Shipping Cost Base * (Undivided × Total)`
        - Where:
          - Width in CM = (Width in inches + 2) × 2.5
          - Undivided = Number of Blinds (Undivided column value)
          - Total = Total (for Shipping column value)

        **3. Final Cost Summary Calculations:**
        - **Product Profit** = `Product Sub-Total * Product Profit %`
        - `Net Profit` = Sum of all `Product Profit` values.
        - `Bill Total` = `Overall Sub-Total` + `Total Estimated Shipping` + `Total Motor Cost`
        """)

if __name__ == "__main__":
    main()